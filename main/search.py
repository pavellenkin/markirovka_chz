import bs4
import requests
from bs4 import BeautifulSoup
import json
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
import difflib
import re
# from main.api.cert import load_cert
# from main.api.true_api import auth



"""
4810737003603

4008177190803

4008177180347

5415236818070

8806314107914

4044688587756 - алекс авто

4607038893985 - косметика LAVR

4036021950280 - косметика

"""


def get_nac_cat(shk):
    url =f"https://xn----7sbabas4ajkhfocclk9d3cvfsa.xn--p1ai/search/?q={shk}&type=goods"
    query_get = requests.get(url)
    html_content = query_get.content.decode()
    soup = BeautifulSoup(html_content, 'lxml')
    try:
        card = soup.find("div", {"class":"catalog__grid-view"})
        href = card.find("a", {"class":"link-gray"})
        result_href = href.get('href')
        result_text =href.text

    except AttributeError:
        return None, None
    return result_href, result_text

def send_link(link):
    identify_item =""
    query_get = requests.get(link)
    html_content = query_get.content.decode()
    # print(html_content)
    soup = BeautifulSoup(html_content, 'lxml')
    card_name = soup.find("div", {"class": "pad-b-20"})
    name_org = card_name.text.replace("\n", "").lstrip().replace("Подписано цифровой подписью ", "")
    card_img = soup.find("div", {"class": "single-product__slider-main__item"})

    img_place = card_img.img['data-hide']
    code_div = soup.find("div", {"class": "pull-right"})
    code = code_div.text
    identify_card = soup.findAll("div", {"class": "tab-section"})

    for element_in in identify_card:

        heading = element_in.find('div', {"class":'tab-section__heading'})
        if heading is not None:
            if 'Идентификация товара' in heading.text:

                for table_element in element_in.table:
                    if table_element:
                        str_info = table_element.text.strip().rstrip().replace("\n\n", ":")
                        str_inden = f"{re.sub(r'\s+', ' ', str_info)}<br>"
                        identify_item += str_inden


            if 'Эксплуатация и условия хранения' in heading.text:
                for table_element in element_in.table:
                    if table_element:
                        str_info = table_element.text.strip().rstrip().replace("\n\n", ":")
                        str_if = f"{re.sub(r'\s+', ' ', str_info)}<br>"
                        identify_item += str_if
    return name_org, img_place, code, identify_item

def search_inn(name):

    url=f"https://datanewton.ru/search.data?query={name}"
    query_get = requests.get(url, )
    html_content = query_get.content.decode()

    res = json.loads(html_content)

    # for i in range(len(res)):
    #     print(res[i])
    index_inn = 'inn'
    index_name = 'name'
    try:
        inn = res[res.index(index_inn) + 1:][0]
    except ValueError:
        inn = "отсутствует"
    try:
        name_company = res[res.index(index_name) + 1:][0]
    except ValueError:
        name_company = ""
    innstring = (
        f"{name_company}<br> <b>"
        f'<a href="#" onclick="'
        f"CopyToClipboard("
        f"{inn}"
        f');"'
        f'title="Копировать" style="text-decoration: none; color: black">'
        f'<h6 id="{inn}" class="m-3 card-text fw-bold">'
        f"{inn}"
        f"</h6></a>"
        f"</b>")
    return name_company, innstring

def search_inn_to_file(name):
    print(name)
    temp_list = []

    file = "list_company.xlsx"
    company_all_obj = load_workbook(filename=file)
    sheet = company_all_obj['Sheet1']
    row_count = sheet.max_row
    # print(row_count)
    # print(x)
    for i in range(2, row_count + 1):
        key_name = sheet.cell(row=i, column=2).value.lower()
        if name in key_name:
            inn = (sheet.cell(row=i, column=1).value
                   .replace('=', '')
                   .replace('"', '')
                   )
            temp_list.append(
                f"{sheet.cell(row=i, column=2).value}<br> <b>"
                f'<a href="#" onclick="'
                f"CopyToClipboard("
                f"{i}"
                f');"'
                
                f'title="Копировать" style="text-decoration: none; color: black">'
                f'<h6 id="{i}" class="m-3 card-text fw-bold">'
                f"{inn}"
                f"</h6></a>"
                f"</b>"
            )
    if temp_list:
        return temp_list
    else:
        return None


def search_inn_to_file_priority_gtin():
    temp_list = []
    gtin = ""
    file = "processed.xlsx"
    inn ="100083608"
    company_all_obj = load_workbook(filename=file)
    sheet = company_all_obj['Sheet1']
    row_count = sheet.max_row
    print(inn)
    for i in range(2, row_count+1):
        try:
            key_inn = sheet.cell(row=i, column=3).value
        except AttributeError:
            key_inn = ""
        if key_inn == inn:
            try:
                gtin =  sheet.cell(row=i, column=1).value
                temp_list.append(gtin)
            except AttributeError:
                gtin = ""
    #print(temp_list)
    for i in range(len(temp_list)):
        print(temp_list[i])





def search_inn_to_file_priority(item):
    file = "processed.xlsx"
    inn =""
    company_all_obj = load_workbook(filename=file)
    sheet = company_all_obj['Sheet1']
    row_count = sheet.max_row
    print(item)
    for i in range(2, row_count+1):
        try:
            key_name = sheet.cell(row=i, column=1).value.lower()[1:]
        except AttributeError:
            key_name = ""
        if item == key_name:
            print(key_name)
            try:
                inn =  sheet.cell(row=i, column=3).value.lower()
            except AttributeError:
                inn = ""
    print(inn)
    if inn:
        return inn
    else:
        return None


def get_nac_cat_all(query):
    temp_dict = {}
    temp_list = []
    url =f"https://xn----7sbabas4ajkhfocclk9d3cvfsa.xn--p1ai/search/getSearchResult?type=goods&query={query}"
    query_get = requests.get(url)
    html_content = query_get.content.decode()
    res = json.loads(html_content)
    # print(res['suggestions'])
    for i in range(len(res['suggestions'])):
        print(res['suggestions'][i]['value'])
        temp_dict['Id']= i
        temp_dict['Value'] = res['suggestions'][i]['value']
        temp_list.append(temp_dict)
        temp_dict = {}
    return temp_list
    soup = BeautifulSoup(html_content, 'lxml')

    # card = soup.findAll("div", {"class":"catalog__grid-view"})
    # print(card)
    # href = card.find("a", {"class":"link-gray"})
    # result_href = href.get('href')
    # result_text =href.text

def identify_tm(tm):
    url = f"https://xn----7sbabas4ajkhfocclk9d3cvfsa.xn--p1ai/search/?q={tm}&type=tm"
    query_get = requests.get(url)
    html_content = query_get.content.decode()
    # print(html_content)
    soup = BeautifulSoup(html_content, 'lxml')
    tag_catalog = soup.find("span", {"class":'catalog__header__items-count'})
    return tag_catalog.text.strip().replace("\n", "")


def get_nac_cat_all_tm(query):
    temp_dict = {}
    temp_list = []
    url = f"https://xn----7sbabas4ajkhfocclk9d3cvfsa.xn--p1ai/search/getSearchResult?type=tm&query={query}"
    query_get = requests.get(url)
    html_content = query_get.content.decode()
    res = json.loads(html_content)
    # print(res['suggestions'])
    for i in range(len(res['suggestions'])):

        identify_tm_search = identify_tm(res['suggestions'][i]['value'])
        if identify_tm_search != "По вашему запросу ничего не найдено":

            temp_dict['Id'] = i
            temp_dict['Value'] = f"<h4><strong>{res['suggestions'][i]['value']}</strong></h4> | {identify_tm_search} |"
            temp_dict['Brand'] = res['suggestions'][i]['value']

            temp_list.append(temp_dict)
            temp_dict = {}
    return temp_list







"""
bitrix_sessid
"""
# def auth_nk():
#     sess = requests.Session()
#     url = "https://markirovka.crpt.ru/bff-elk/v1/united-auth/login"
#     header = {
#             "Accept" : "application/json, text/plain, */*",
#              "Content-Type" : "application/json; charset=UTF-8"
#          }
#     tok = auth()
#     print(tok)
#     data = {"keep":{
#         "data": "MIIbJwYJKoZIhvcNAQcCoIIbGDCCGxQCAQExDjAMBggqhQMHAQECAgUAMBkGCSqGSIb3DQEHAaAMBAo3NzEyMDM1NzI5oIIWeTCCBVEwggT+oAMCAQICEQCVH6NHfGEEOq36hYYngjRCMAoGCCqFAwcBAQMCMIIBOzEhMB8GCSqGSIb3DQEJARYSZGl0QGRpZ2l0YWwuZ292LnJ1MQswCQYDVQQGEwJSVTEYMBYGA1UECAwPNzcg0JzQvtGB0LrQstCwMRkwFwYDVQQHDBDQsy4g0JzQvtGB0LrQstCwMVMwUQYDVQQJDErQn9GA0LXRgdC90LXQvdGB0LrQsNGPINC90LDQsdC10YDQtdC20L3QsNGPLCDQtNC+0LwgMTAsINGB0YLRgNC+0LXQvdC40LUgMjEmMCQGA1UECgwd0JzQuNC90YbQuNGE0YDRiyDQoNC+0YHRgdC40LgxGDAWBgUqhQNkARINMTA0NzcwMjAyNjcwMTEVMBMGBSqFA2QEEgo3NzEwNDc0Mzc1MSYwJAYDVQQDDB3QnNC40L3RhtC40YTRgNGLINCg0L7RgdGB0LjQuDAeFw0yMjAxMDgxMzMyMzlaFw00MDAxMDgxMzMyMzlaMIIBOzEhMB8GCSqGSIb3DQEJARYSZGl0QGRpZ2l0YWwuZ292LnJ1MQswCQYDVQQGEwJSVTEYMBYGA1UECAwPNzcg0JzQvtGB0LrQstCwMRkwFwYDVQQHDBDQsy4g0JzQvtGB0LrQstCwMVMwUQYDVQQJDErQn9GA0LXRgdC90LXQvdGB0LrQsNGPINC90LDQsdC10YDQtdC20L3QsNGPLCDQtNC+0LwgMTAsINGB0YLRgNC+0LXQvdC40LUgMjEmMCQGA1UECgwd0JzQuNC90YbQuNGE0YDRiyDQoNC+0YHRgdC40LgxGDAWBgUqhQNkARINMTA0NzcwMjAyNjcwMTEVMBMGBSqFA2QEEgo3NzEwNDc0Mzc1MSYwJAYDVQQDDB3QnNC40L3RhtC40YTRgNGLINCg0L7RgdGB0LjQuDBmMB8GCCqFAwcBAQEBMBMGByqFAwICIwEGCCqFAwcBAQICA0MABEBaSmukHWuPC4xav3a89jNu3xarv4N/j68a4PZRPij83W70R8LjrW4ZSfdqIJkvou5oQwxj7FobT1XblfSm6kCOo4IB0DCCAcwwgfUGBSqFA2RwBIHrMIHoDDTQn9CQ0JrQnCDCq9Ca0YDQuNC/0YLQvtCf0YDQviBIU03CuyDQstC10YDRgdC40LggMi4wDEPQn9CQ0JogwqvQk9C+0LvQvtCy0L3QvtC5INGD0LTQvtGB0YLQvtCy0LXRgNGP0Y7RidC40Lkg0YbQtdC90YLRgMK7DDXQl9Cw0LrQu9GO0YfQtdC90LjQtSDihJYgMTQ5LzMvMi8yLzIzINC+0YIgMDIuMDMuMjAxOAw00JfQsNC60LvRjtGH0LXQvdC40LUg4oSWIDE0OS83LzYtNDQ5INC+0YIgMzAuMTIuMjAyMTA/BgUqhQNkbwQ2DDTQn9CQ0JrQnCDCq9Ca0YDQuNC/0YLQvtCf0YDQviBIU03CuyDQstC10YDRgdC40LggMi4wMAwGBSqFA2RyBAMCAQAwQwYDVR0gBDwwOjAIBgYqhQNkcQEwCAYGKoUDZHECMAgGBiqFA2RxAzAIBgYqhQNkcQQwCAYGKoUDZHEFMAYGBFUdIAAwDgYDVR0PAQH/BAQDAgEGMA8GA1UdEwEB/wQFMAMBAf8wHQYDVR0OBBYEFMkTWLFMp2I6ftI/PKbnFHydcKOGMAoGCCqFAwcBAQMCA0EAgkl4SAoN0mf00wrUSaQRblxkstz3mUxHjsHThyL9LVojdqtJj0cwf6Oq/TF4xqnc32BrQq6tZeRbF2TiOXP9kjCCCBIwgge/oAMCAQICCwDntF1kAAAAAArQMAoGCCqFAwcBAQMCMIIBOzEhMB8GCSqGSIb3DQEJARYSZGl0QGRpZ2l0YWwuZ292LnJ1MQswCQYDVQQGEwJSVTEYMBYGA1UECAwPNzcg0JzQvtGB0LrQstCwMRkwFwYDVQQHDBDQsy4g0JzQvtGB0LrQstCwMVMwUQYDVQQJDErQn9GA0LXRgdC90LXQvdGB0LrQsNGPINC90LDQsdC10YDQtdC20L3QsNGPLCDQtNC+0LwgMTAsINGB0YLRgNC+0LXQvdC40LUgMjEmMCQGA1UECgwd0JzQuNC90YbQuNGE0YDRiyDQoNC+0YHRgdC40LgxGDAWBgUqhQNkARINMTA0NzcwMjAyNjcwMTEVMBMGBSqFA2QEEgo3NzEwNDc0Mzc1MSYwJAYDVQQDDB3QnNC40L3RhtC40YTRgNGLINCg0L7RgdGB0LjQuDAeFw0yNTAzMjUxMDA4MDNaFw0zOTAzMjUxMDA4MDNaMIIBhDEVMBMGBSqFA2QEEgo2NjYzMDAzMTI3MR4wHAYJKoZIhvcNAQkBFg9jYUBza2Jrb250dXIucnUxGDAWBgUqhQNkARINMTAyNjYwNTYwNjYyMDELMAkGA1UEBhMCUlUxMzAxBgNVBAgMKjY2INCh0LLQtdGA0LTQu9C+0LLRgdC60LDRjyDQvtCx0LvQsNGB0YLRjDEhMB8GA1UEBwwY0JXQutCw0YLQtdGA0LjQvdCx0YPRgNCzMUQwQgYDVQQJDDvRg9C70LjRhtCwINCd0LDRgNC+0LTQvdC+0Lkg0LLQvtC70LgsINGB0YLRgNC+0LXQvdC40LUgMTnQkDEwMC4GA1UECwwn0KPQtNC+0YHRgtC+0LLQtdGA0Y/RjtGJ0LjQuSDRhtC10L3RgtGAMSkwJwYDVQQKDCDQkNCeICLQn9CkICLQodCa0JEg0JrQvtC90YLRg9GAIjEpMCcGA1UEAwwg0JDQniAi0J/QpCAi0KHQmtCRINCa0L7QvdGC0YPRgCIwZjAfBggqhQMHAQEBATATBgcqhQMCAiMBBggqhQMHAQECAgNDAARAQx8l4FmIk8lRloOCxcne7VoPT2S8nFoOpd5jKiLN3kqG2BsXPZEmYQX4NYET6SE+iPud5DZBViBdOipuGEMvUaOCBE4wggRKMAsGA1UdDwQEAwIBhjAdBgNVHQ4EFgQU1TJPisbTBIci0gpnPK8jfypS2RYwEgYDVR0TAQH/BAgwBgEB/wIBADAlBgNVHSAEHjAcMAgGBiqFA2RxATAIBgYqhQNkcQIwBgYEVR0gADArBgNVHRAEJDAigA8yMDI1MDMxNzEwMDQyMVqBDzIwMjgwMzE3MTAwNDIxWjBUBgUqhQNkbwRLDEki0JrRgNC40L/RgtC+0J/RgNC+IENTUCIgKNCy0LXRgNGB0LjRjyA0LjApICjQuNGB0L/QvtC70L3QtdC90LjQtSAyLUJhc2UpMBQGCSsGAQQBgjcUAgQHDAVTdWJDQTASBgkrBgEEAYI3FQEEBQIDBgAGMIIBfQYDVR0jBIIBdDCCAXCAFMkTWLFMp2I6ftI/PKbnFHydcKOGoYIBQ6SCAT8wggE7MSEwHwYJKoZIhvcNAQkBFhJkaXRAZGlnaXRhbC5nb3YucnUxCzAJBgNVBAYTAlJVMRgwFgYDVQQIDA83NyDQnNC+0YHQutCy0LAxGTAXBgNVBAcMENCzLiDQnNC+0YHQutCy0LAxUzBRBgNVBAkMStCf0YDQtdGB0L3QtdC90YHQutCw0Y8g0L3QsNCx0LXRgNC10LbQvdCw0Y8sINC00L7QvCAxMCwg0YHRgtGA0L7QtdC90LjQtSAyMSYwJAYDVQQKDB3QnNC40L3RhtC40YTRgNGLINCg0L7RgdGB0LjQuDEYMBYGBSqFA2QBEg0xMDQ3NzAyMDI2NzAxMRUwEwYFKoUDZAQSCjc3MTA0NzQzNzUxJjAkBgNVBAMMHdCc0LjQvdGG0LjRhNGA0Ysg0KDQvtGB0YHQuNC4ghEAlR+jR3xhBDqt+oWGJ4I0QjBoBgNVHR8EYTBfMC2gK6AphidodHRwOi8vY3JsLmdvc3VzbHVnaS5ydS9jZHAvZ3VjMjAyMi5jcmwwLqAsoCqGKGh0dHA6Ly9jcmwyLmdvc3VzbHVnaS5ydS9jZHAvZ3VjMjAyMi5jcmwwQwYIKwYBBQUHAQEENzA1MDMGCCsGAQUFBzAChidodHRwOi8vY3JsLmdvc3VzbHVnaS5ydS9jZHAvZ3VjMjAyMi5jcnQwgfUGBSqFA2RwBIHrMIHoDDTQn9CQ0JrQnCDCq9Ca0YDQuNC/0YLQvtCf0YDQviBIU03CuyDQstC10YDRgdC40LggMi4wDEPQn9CQ0JogwqvQk9C+0LvQvtCy0L3QvtC5INGD0LTQvtGB0YLQvtCy0LXRgNGP0Y7RidC40Lkg0YbQtdC90YLRgMK7DDXQl9Cw0LrQu9GO0YfQtdC90LjQtSDihJYgMTQ5LzMvMi8yLzIzINC+0YIgMDIuMDMuMjAxOAw00JfQsNC60LvRjtGH0LXQvdC40LUg4oSWIDE0OS83LzYtNDQ5INC+0YIgMzAuMTIuMjAyMTAMBgUqhQNkcgQDAgEBMAoGCCqFAwcBAQMCA0EA1f+oeVB3hMom4qSwsuY3ryrEA2bTKjCfSpco0caT4YVmlvWK2nwiNYA6f62AugGadbyXtI2VgBwKBuZHCs91vjCCCQowggi3oAMCAQICEQb+2dAALLPbm06o6EuLz7WxMAoGCCqFAwcBAQMCMIIBhDEVMBMGBSqFA2QEEgo2NjYzMDAzMTI3MR4wHAYJKoZIhvcNAQkBFg9jYUBza2Jrb250dXIucnUxGDAWBgUqhQNkARINMTAyNjYwNTYwNjYyMDELMAkGA1UEBhMCUlUxMzAxBgNVBAgMKjY2INCh0LLQtdGA0LTQu9C+0LLRgdC60LDRjyDQvtCx0LvQsNGB0YLRjDEhMB8GA1UEBwwY0JXQutCw0YLQtdGA0LjQvdCx0YPRgNCzMUQwQgYDVQQJDDvRg9C70LjRhtCwINCd0LDRgNC+0LTQvdC+0Lkg0LLQvtC70LgsINGB0YLRgNC+0LXQvdC40LUgMTnQkDEwMC4GA1UECwwn0KPQtNC+0YHRgtC+0LLQtdGA0Y/RjtGJ0LjQuSDRhtC10L3RgtGAMSkwJwYDVQQKDCDQkNCeICLQn9CkICLQodCa0JEg0JrQvtC90YLRg9GAIjEpMCcGA1UEAwwg0JDQniAi0J/QpCAi0KHQmtCRINCa0L7QvdGC0YPRgCIwHhcNMjUwODAxMTIzNTI1WhcNMjYwODAxMTI0MDI1WjCB8DEpMCcGCSqGSIb3DQEJARYaZWdvci5nb3JiYWNoZXZAYXV0b2V1cm8ucnUxGjAYBggqhQMDgQMBARIMNDEwMTE3OTUxNTU5MRYwFAYFKoUDZAMSCzE2MTU2MjM4NTYxMQswCQYDVQQGEwJSVTEqMCgGA1UEKgwh0JXQs9C+0YAg0JLQu9Cw0LTQuNC80LjRgNC+0LLQuNGHMRkwFwYDVQQEDBDQk9C+0YDQsdCw0YfQtdCyMTswOQYDVQQDDDLQk9C+0YDQsdCw0YfQtdCyINCV0LPQvtGAINCS0LvQsNC00LjQvNC40YDQvtCy0LjRhzBmMB8GCCqFAwcBAQEBMBMGByqFAwICJAAGCCqFAwcBAQICA0MABEAyRv2Wrubvw0gyxfWYbTvegOgSKG9fgVmW12H1i9V94PUrpJgHVpj5PPeXO18Vl59rBMuptmEBbOqgCkjWGlHno4IFjDCCBYgwDAYFKoUDZHIEAwIBADAOBgNVHQ8BAf8EBAMCBPAwJQYDVR0RBB4wHIEaZWdvci5nb3JiYWNoZXZAYXV0b2V1cm8ucnUwEwYDVR0gBAwwCjAIBgYqhQNkcQEwLwYDVR0lBCgwJgYIKwYBBQUHAwIGByqFAwICIgYGCCsGAQUFBwMEBgcqhQMDBwgBMIHWBggrBgEFBQcBAQSByTCBxjA1BggrBgEFBQcwAYYpaHR0cDovL3BraS5za2Jrb250dXIucnUvb2NzcHFjYTYvb2NzcC5zcmYwRQYIKwYBBQUHMAKGOWh0dHA6Ly9jZHAuc2tia29udHVyLnJ1L2NlcnRpZmljYXRlcy9za2Jrb250dXItcS0yMDI1LmNydDBGBggrBgEFBQcwAoY6aHR0cDovL2NkcDIuc2tia29udHVyLnJ1L2NlcnRpZmljYXRlcy9za2Jrb250dXItcS0yMDI1LmNydDArBgNVHRAEJDAigA8yMDI1MDgwMTEyMzUyNFqBDzIwMjYwODAxMTI0MDI0WjCCATMGBSqFA2RwBIIBKDCCASQMKyLQmtGA0LjQv9GC0L7Qn9GA0L4gQ1NQIiAo0LLQtdGA0YHQuNGPIDQuMCkMUyLQo9C00L7RgdGC0L7QstC10YDRj9GO0YnQuNC5INGG0LXQvdGC0YAgItCa0YDQuNC/0YLQvtCf0YDQviDQo9CmIiDQstC10YDRgdC40LggMi4wDE/QodC10YDRgtC40YTQuNC60LDRgiDRgdC+0L7RgtCy0LXRgtGB0YLQstC40Y8g4oSWINCh0KQvMTI0LTQ3MTcg0L7RgiAxNS4wMS4yMDI0DE/QodC10YDRgtC40YTQuNC60LDRgiDRgdC+0L7RgtCy0LXRgtGB0YLQstC40Y8g4oSWINCh0KQvMTI4LTUyMTUg0L7RgiAxMC4wNy4yMDI1MCMGBSqFA2RvBBoMGCLQmtGA0LjQv9GC0L7Qn9GA0L4gQ1NQIjB6BgNVHR8EczBxMDagNKAyhjBodHRwOi8vY2RwLnNrYmtvbnR1ci5ydS9jZHAvc2tia29udHVyLXEtMjAyNS5jcmwwN6A1oDOGMWh0dHA6Ly9jZHAyLnNrYmtvbnR1ci5ydS9jZHAvc2tia29udHVyLXEtMjAyNS5jcmwwgYIGByqFAwICMQIEdzB1MGUWQGh0dHBzOi8vY2Eua29udHVyLnJ1L2Fib3V0L2RvY3VtZW50cy9jcnlwdG9wcm8tbGljZW5zZS1xdWFsaWZpZWQMHdCh0JrQkSDQmtC+0L3RgtGD0YAg0Lgg0JTQl9CeAwIF4AQMN8kkT25fX8UGHBdCMIIBdwYDVR0jBIIBbjCCAWqAFNUyT4rG0wSHItIKZzyvI38qUtkWoYIBQ6SCAT8wggE7MSEwHwYJKoZIhvcNAQkBFhJkaXRAZGlnaXRhbC5nb3YucnUxCzAJBgNVBAYTAlJVMRgwFgYDVQQIDA83NyDQnNC+0YHQutCy0LAxGTAXBgNVBAcMENCzLiDQnNC+0YHQutCy0LAxUzBRBgNVBAkMStCf0YDQtdGB0L3QtdC90YHQutCw0Y8g0L3QsNCx0LXRgNC10LbQvdCw0Y8sINC00L7QvCAxMCwg0YHRgtGA0L7QtdC90LjQtSAyMSYwJAYDVQQKDB3QnNC40L3RhtC40YTRgNGLINCg0L7RgdGB0LjQuDEYMBYGBSqFA2QBEg0xMDQ3NzAyMDI2NzAxMRUwEwYFKoUDZAQSCjc3MTA0NzQzNzUxJjAkBgNVBAMMHdCc0LjQvdGG0LjRhNGA0Ysg0KDQvtGB0YHQuNC4ggsA57RdZAAAAAAK0DAdBgNVHQ4EFgQUHE0rDci2You9mOrI/U5KHF/kr+YwCgYIKoUDBwEBAwIDQQCTw+pR+1MZMjJcphb5oj3VKeiokxsdNr8V3OnZBN5fatdo0HovAE7HESI2Puje7C42IePGoB2TMDR1S8HzF0QLMYIEZTCCBGECAQEwggGbMIIBhDEVMBMGBSqFA2QEEgo2NjYzMDAzMTI3MR4wHAYJKoZIhvcNAQkBFg9jYUBza2Jrb250dXIucnUxGDAWBgUqhQNkARINMTAyNjYwNTYwNjYyMDELMAkGA1UEBhMCUlUxMzAxBgNVBAgMKjY2INCh0LLQtdGA0LTQu9C+0LLRgdC60LDRjyDQvtCx0LvQsNGB0YLRjDEhMB8GA1UEBwwY0JXQutCw0YLQtdGA0LjQvdCx0YPRgNCzMUQwQgYDVQQJDDvRg9C70LjRhtCwINCd0LDRgNC+0LTQvdC+0Lkg0LLQvtC70LgsINGB0YLRgNC+0LXQvdC40LUgMTnQkDEwMC4GA1UECwwn0KPQtNC+0YHRgtC+0LLQtdGA0Y/RjtGJ0LjQuSDRhtC10L3RgtGAMSkwJwYDVQQKDCDQkNCeICLQn9CkICLQodCa0JEg0JrQvtC90YLRg9GAIjEpMCcGA1UEAwwg0JDQniAi0J/QpCAi0KHQmtCRINCa0L7QvdGC0YPRgCICEQb+2dAALLPbm06o6EuLz7WxMAwGCCqFAwcBAQICBQCgggJfMBgGCSqGSIb3DQEJAzELBgkqhkiG9w0BBwEwHAYJKoZIhvcNAQkFMQ8XDTI1MDkyMzA4MjAzN1owLwYJKoZIhvcNAQkEMSIEIN6dimnRhALAy3zAKzlNrOiCtRiKVbEPfgslRAK5yFo3MIIB8gYLKoZIhvcNAQkQAi8xggHhMIIB3TCCAdkwggHVMAoGCCqFAwcBAQICBCAfZcn/KCRw0vccorXU2vwfhVePKWocr+bqA/ZTOPTu/zCCAaMwggGMpIIBiDCCAYQxFTATBgUqhQNkBBIKNjY2MzAwMzEyNzEeMBwGCSqGSIb3DQEJARYPY2FAc2tia29udHVyLnJ1MRgwFgYFKoUDZAESDTEwMjY2MDU2MDY2MjAxCzAJBgNVBAYTAlJVMTMwMQYDVQQIDCo2NiDQodCy0LXRgNC00LvQvtCy0YHQutCw0Y8g0L7QsdC70LDRgdGC0YwxITAfBgNVBAcMGNCV0LrQsNGC0LXRgNC40L3QsdGD0YDQszFEMEIGA1UECQw70YPQu9C40YbQsCDQndCw0YDQvtC00L3QvtC5INCy0L7Qu9C4LCDRgdGC0YDQvtC10L3QuNC1IDE50JAxMDAuBgNVBAsMJ9Cj0LTQvtGB0YLQvtCy0LXRgNGP0Y7RidC40Lkg0YbQtdC90YLRgDEpMCcGA1UECgwg0JDQniAi0J/QpCAi0KHQmtCRINCa0L7QvdGC0YPRgCIxKTAnBgNVBAMMINCQ0J4gItCf0KQgItCh0JrQkSDQmtC+0L3RgtGD0YAiAhEG/tnQACyz25tOqOhLi8+1sTAKBggqhQMHAQEBAQRA61JlqsCPQZv82hyX5lhlhFKfNeKWC+e6/WEEbdCaTGcB8BqqNIwMOba3+8kUC7WEidkbhNf81EQexXD+68rhuw==",
#         "mrdToken":"true"
#     }}
#     get_site = sess.post(url, json=data, headers=header)
#     html_content = get_site.content.decode()
#     print(html_content)
#     # res = json.loads(html_content)
#     # signed_data = load_cert(res['data'])
#     # # print(signed_data)
#     # inn = "7712035729"
#     # data_cert = {
#     #     "fingerprint" : "",
#     #     "signature": signed_data,
#     #     "inn": "7712035729",
#     #
#     # }
#     # header = {
#     #     "Accept" : "application/json, text/plain, */*",
#     #     "Content-Type" : "application/json; charset=UTF-8"
#     # }
#     # url_cert_login = "https://xn--j1ab.xn----7sbabas4ajkhfocclk9d3cvfsa.xn--p1ai/rest/certlogin"
#     # cert_login = sess.post(url_cert_login, json=data_cert, headers=header)
#     # cert_login_html_content = cert_login.content.decode()
#     # res_cert = json.loads(cert_login_html_content)
#     # print(res_cert)


# def check_tnved():
#     token = ""
#     url_token = "https://xn--80ajghhoc2aj1c8b.xn--p1ai/checking_codes"
#     sess = requests.Session()
#     url = "https://xn--80ajghhoc2aj1c8b.xn--p1ai/bitrix/services/main/ajax.php?mode=class&c=dev%3AcodeSearch&action=getByTnVed"
#
#     get_site = sess.get(url_token)
#     html_content = get_site.content.decode()
#     soup = BeautifulSoup(html_content, 'lxml')
#     tag_script = soup.findAll("script")
#     for i in tag_script:
#         if "bitrix_sessid" in i.text:
#             search_token = i.text.split("bitrix_sessid")
#             token = search_token[-1].replace("'});", "").replace("':'", "")
#     print(token)
#     headers = {
#         "mode" :"class",
#         "c": "dev:codeSearch",
#         "action":"getByTnVed"
#     }
#     data = {
#         "inputValue": "3402500000",
#         "SITE_ID": "s1",
#         "sessid": token
#     }
#
#     query_post = sess.post(url, json=data, headers=headers)
#     print(query_post.content.decode())




